#!/usr/bin/env python3
"""
FlorLogic — generador de la semilla de la demo de captura.

Lee la plantilla digitalizada del cliente (formatos reales de Buenavista) y produce
public/seed.json: catalogo de bloques, camas, variedades y siembras historicas.

No inventa datos. Todo lo que no se puede derivar del archivo queda marcado
como desconocido y se reporta al final de la corrida.

Uso:
    python scripts/generar_seed.py "<ruta al .xlsx>" [-o public/seed.json]
"""
from __future__ import annotations

import argparse
import json
import re
import statistics
import sys
import unicodedata
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path

try:
    import openpyxl
except ImportError:  # pragma: no cover
    sys.exit("Falta openpyxl.  pip install openpyxl")

# --------------------------------------------------------------------------------------
# Configuracion de las dos plantillas reales encontradas en el archivo del cliente.
# Ver PLAN_DEMO_CAPTURA.md seccion 2.4: NO son el mismo formato.
# --------------------------------------------------------------------------------------

HOJAS_NOVEDAD = ("Cremón", "Cremon", "Matsomoto")
HOJA_PROGRAMA = ("Programa Siembras", "Programa de Siembras")

PLANTILLAS = [
    {
        "id": "novedad_siembra",
        "nombre": "Novedades de siembra",
        "campos": [
            {"id": "fecha", "etiqueta": "Fecha", "tipo": "fecha", "obligatorio": True},
            {"id": "bloque", "etiqueta": "Bloque", "tipo": "catalogo", "obligatorio": True},
            {"id": "cama", "etiqueta": "Cama", "tipo": "catalogo", "obligatorio": True},
            {"id": "variedad", "etiqueta": "Variedad", "tipo": "catalogo", "obligatorio": True},
            {"id": "lineas", "etiqueta": "# Líneas", "tipo": "entero", "obligatorio": True},
            {"id": "cantidad", "etiqueta": "Cantidad", "tipo": "entero", "obligatorio": True},
            {"id": "obse", "etiqueta": "OBSE", "tipo": "desconocido", "obligatorio": False},
        ],
    },
    {
        "id": "programa_siembra",
        "nombre": "Programa de siembras",
        "campos": [
            {"id": "fechaSalidaCf", "etiqueta": "Fecha salida CF", "tipo": "fecha", "obligatorio": True},
            {"id": "fechaSiembra", "etiqueta": "Fecha de siembra", "tipo": "fecha", "obligatorio": True},
            {"id": "longPrebrotado", "etiqueta": "Long. prebrotado", "tipo": "entero", "obligatorio": False},
            {"id": "bloque", "etiqueta": "Bloque", "tipo": "catalogo", "obligatorio": True},
            {"id": "cama", "etiqueta": "Cama", "tipo": "catalogo", "obligatorio": True},
            {"id": "variedad", "etiqueta": "Variedad", "tipo": "catalogo", "obligatorio": True},
            {"id": "lote", "etiqueta": "Lote", "tipo": "texto", "obligatorio": True},
            {"id": "calibre", "etiqueta": "Calibre", "tipo": "texto", "obligatorio": False},
            {"id": "proveedor", "etiqueta": "Proveedor", "tipo": "catalogo", "obligatorio": True},
            {"id": "contenedor", "etiqueta": "Contenedor", "tipo": "texto", "obligatorio": False},
            {"id": "observaciones", "etiqueta": "Observaciones", "tipo": "desconocido", "obligatorio": False},
        ],
    },
]


def slug(texto: str) -> str:
    t = unicodedata.normalize("NFKD", str(texto)).encode("ascii", "ignore").decode()
    t = re.sub(r"[^a-zA-Z0-9]+", "-", t).strip("-").lower()
    return t or "sin-nombre"


def limpiar(valor) -> str:
    if valor is None:
        return ""
    if isinstance(valor, datetime):
        return valor.date().isoformat()
    return re.sub(r"\s+", " ", str(valor)).strip()


def normalizar_fecha(valor) -> str:
    """El formato real escribe '23 06 26'. Se conserva el original y se deriva ISO."""
    crudo = limpiar(valor)
    m = re.match(r"^(\d{1,2})\s+(\d{1,2})\s+(\d{2,4})$", crudo)
    if not m:
        return crudo
    d, mes, a = (int(x) for x in m.groups())
    a = a + 2000 if a < 100 else a
    try:
        return f"{a:04d}-{mes:02d}-{d:02d}"
    except ValueError:
        return crudo


def entero(valor):
    try:
        return int(round(float(valor)))
    except (TypeError, ValueError):
        return None


class Semilla:
    def __init__(self, origen: Path):
        self.origen = origen
        self.bloques: dict[str, dict] = {}
        self.camas: dict[str, dict] = {}
        self.variedades: dict[str, dict] = {}
        self.siembras: list[dict] = []
        self.avisos: list[str] = []

    # -- catalogo -----------------------------------------------------------------
    def bloque(self, codigo: str, tipo_flor: str) -> str:
        bid = f"blq-{slug(codigo)}"
        b = self.bloques.setdefault(
            bid, {"id": bid, "fincaId": "fnc-buenavista", "codigo": codigo, "tiposFlor": []}
        )
        if tipo_flor and tipo_flor not in b["tiposFlor"]:
            b["tiposFlor"].append(tipo_flor)
        return bid

    def cama(self, bloque_id: str, codigo: str) -> str:
        cid = f"{bloque_id}-cma-{slug(codigo)}"
        self.camas.setdefault(cid, {"id": cid, "bloqueId": bloque_id, "codigo": codigo})
        return cid

    def variedad(self, nombre: str, tipo_flor: str) -> str:
        vid = f"var-{slug(nombre)}"
        v = self.variedades.setdefault(
            vid,
            {"id": vid, "nombre": nombre, "tipoFlor": tipo_flor,
             "plantasPorLinea": None, "confianza": "sin-datos", "razonesObservadas": []},
        )
        return vid

    # -- lectura ------------------------------------------------------------------
    def leer_novedades(self, ws) -> None:
        tipo_flor = limpiar(ws.cell(row=1, column=1).value) or ws.title
        for fila in ws.iter_rows(min_row=4, values_only=False):
            v = [c.value for c in fila]
            if not limpiar(v[2]):
                continue
            cod_bloque = limpiar(v[1])
            cod_cama = limpiar(v[2])
            nombre_var = limpiar(v[3])
            lineas = entero(v[4])
            cantidad = entero(v[5])
            bid = self.bloque(cod_bloque, tipo_flor)
            cid = self.cama(bid, cod_cama)
            vid = self.variedad(nombre_var, tipo_flor)
            razon = round(cantidad / lineas, 3) if lineas else None
            if razon:
                self.variedades[vid]["razonesObservadas"].append(razon)
            self.siembras.append({
                "id": f"sbr-{slug(ws.title)}-{fila[0].row}",
                "plantillaId": "novedad_siembra",
                "hojaOrigen": ws.title,
                "filaOrigen": fila[0].row,
                "fecha": normalizar_fecha(v[0]),
                "fechaCruda": limpiar(v[0]),
                "bloqueId": bid, "camaId": cid, "variedadId": vid,
                "lineas": lineas, "cantidad": cantidad,
                "razonPlantasPorLinea": razon,
                "obse": limpiar(v[6]),
            })

    def leer_programa(self, ws) -> None:
        for fila in ws.iter_rows(min_row=4, values_only=False):
            v = [c.value for c in fila]
            if not limpiar(v[4]):
                continue
            bid = self.bloque(limpiar(v[3]), "")
            cid = self.cama(bid, limpiar(v[4]))
            vid = self.variedad(limpiar(v[5]), "")
            self.siembras.append({
                "id": f"sbr-programa-{fila[0].row}",
                "plantillaId": "programa_siembra",
                "hojaOrigen": ws.title,
                "filaOrigen": fila[0].row,
                "fecha": normalizar_fecha(v[1]),
                "fechaCruda": limpiar(v[1]),
                "fechaSalidaCf": normalizar_fecha(v[0]),
                "longPrebrotado": entero(v[2]),
                "bloqueId": bid, "camaId": cid, "variedadId": vid,
                "lineas": None, "cantidad": None, "razonPlantasPorLinea": None,
                "lote": limpiar(v[6]), "calibre": limpiar(v[7]),
                "proveedor": limpiar(v[8]), "contenedor": limpiar(v[9]),
                "observaciones": limpiar(v[10]),
            })

    # -- derivaciones -------------------------------------------------------------
    def derivar_plantas_por_linea(self) -> None:
        """La razon dominante por tipo de flor es la regla; la de la variedad la refina.

        Ver PLAN_DEMO_CAPTURA.md 2.3: Cremon ~19, Matsomoto ~15.
        """
        por_flor = defaultdict(list)
        for s in self.siembras:
            if s["razonPlantasPorLinea"]:
                flor = self.variedades[s["variedadId"]]["tipoFlor"]
                por_flor[flor].append(s["razonPlantasPorLinea"])

        modo_flor = {}
        for flor, razones in por_flor.items():
            modo_flor[flor] = Counter(round(r) for r in razones).most_common(1)[0][0]

        for v in self.variedades.values():
            razones = v["razonesObservadas"]
            if not razones:
                v["plantasPorLinea"] = modo_flor.get(v["tipoFlor"])
                v["confianza"] = "heredada-del-tipo-de-flor" if v["plantasPorLinea"] else "sin-datos"
                continue
            conteo = Counter(round(r) for r in razones)
            valor, veces = conteo.most_common(1)[0]
            v["plantasPorLinea"] = valor
            if len(conteo) == 1:
                v["confianza"] = "alta"
            elif veces >= max(2, len(razones) - 1):
                v["confianza"] = "media"
            else:
                v["confianza"] = "en-disputa"
                self.avisos.append(
                    f"Variedad '{v['nombre']}': {len(conteo)} razones distintas "
                    f"({sorted(set(razones))}) — no hay valor confiable de plantas por línea."
                )

    def detectar_anomalias(self) -> list[dict]:
        """Aplica la primera regla dura sobre el historico real del cliente."""
        anomalias = []

        # A · la cantidad no cuadra con lineas x plantas por linea
        for s in self.siembras:
            if not s["razonPlantasPorLinea"]:
                continue
            v = self.variedades[s["variedadId"]]
            esperado = v["plantasPorLinea"]
            if not esperado:
                continue
            desvio = abs(s["razonPlantasPorLinea"] - esperado) / esperado
            if desvio > 0.02:
                anomalias.append({
                    "tipo": "razon-plantas-por-linea",
                    "severidad": "revisar",
                    "siembraId": s["id"],
                    "hoja": s["hojaOrigen"], "fila": s["filaOrigen"],
                    "detalle": (
                        f"{v['nombre']}: {s['cantidad']} / {s['lineas']} = "
                        f"{s['razonPlantasPorLinea']} plantas por línea, "
                        f"pero la variedad va en {esperado}."
                    ),
                })

        # B · variedades cuyo nombre difiere en una letra: probable error de digitacion
        nombres = [v["nombre"] for v in self.variedades.values() if v["nombre"]]
        for i, a in enumerate(nombres):
            for b in nombres[i + 1:]:
                if a.lower() != b.lower() and len(a) == len(b) and \
                        sum(x != y for x, y in zip(a.lower(), b.lower())) == 1:
                    anomalias.append({
                        "tipo": "variedad-casi-identica",
                        "severidad": "revisar",
                        "detalle": f"'{a}' y '{b}' difieren en una sola letra. ¿Son la misma variedad?",
                    })

        # C · camas con mas de una variedad el mismo dia  ->  NO es error: son secciones (DEC-14)
        agrupado = defaultdict(list)
        for s in self.siembras:
            if s["plantillaId"] == "novedad_siembra":
                agrupado[(s["camaId"], s["fecha"])].append(s)
        for (cama_id, fecha), items in agrupado.items():
            if len(items) > 1:
                anomalias.append({
                    "tipo": "cama-dividida",
                    "severidad": "informativo",
                    "detalle": (
                        f"Cama {self.camas[cama_id]['codigo']} del bloque "
                        f"{self.bloques[self.camas[cama_id]['bloqueId']]['codigo']} el {fecha}: "
                        f"{len(items)} secciones ({', '.join(self.variedades[i['variedadId']]['nombre'] for i in items)})."
                    ),
                })
        return anomalias

    def construir(self) -> dict:
        wb = openpyxl.load_workbook(self.origen, data_only=True)
        for ws in wb.worksheets:
            if ws.title in HOJAS_NOVEDAD:
                self.leer_novedades(ws)
            elif ws.title in HOJA_PROGRAMA:
                self.leer_programa(ws)
            else:
                self.avisos.append(f"Hoja '{ws.title}' ignorada: no coincide con ninguna plantilla conocida.")

        self.derivar_plantas_por_linea()
        anomalias = self.detectar_anomalias()

        for v in self.variedades.values():
            v["razonesObservadas"] = sorted(set(v["razonesObservadas"]))

        obse = sorted({s["obse"] for s in self.siembras if s.get("obse")})
        if obse:
            self.avisos.append(
                f"Columna OBSE: valores distintos observados {obse}. "
                "Significado desconocido — pregunta abierta de la sesión con el cliente."
            )

        return {
            "version": "seed.v1",
            "generado": datetime.now(timezone.utc).isoformat(timespec="seconds"),
            "origen": self.origen.name,
            "advertencia": (
                "Datos reales del cliente. Semilla de demostración: no es un catálogo validado. "
                "Ver PLAN_DEMO_CAPTURA.md §2."
            ),
            "empresa": {"id": "emp-demo", "nombre": "Empresa demo"},
            "fincas": [{"id": "fnc-buenavista", "empresaId": "emp-demo", "nombre": "Buenavista"}],
            "bloques": sorted(self.bloques.values(), key=lambda b: (len(b["codigo"]), b["codigo"])),
            "camas": sorted(self.camas.values(), key=lambda c: (c["bloqueId"], len(c["codigo"]), c["codigo"])),
            "variedades": sorted(self.variedades.values(), key=lambda v: v["nombre"]),
            "plantillas": PLANTILLAS,
            "siembras": self.siembras,
            "anomalias": anomalias,
            "avisos": self.avisos,
        }


def main() -> int:
    p = argparse.ArgumentParser(description="Genera public/seed.json desde la plantilla del cliente.")
    p.add_argument("xlsx", type=Path, help="Ruta a la plantilla digitalizada (.xlsx)")
    p.add_argument("-o", "--salida", type=Path, default=Path("public/seed.json"))
    args = p.parse_args()

    if not args.xlsx.exists():
        return print(f"No existe: {args.xlsx}") or 1

    semilla = Semilla(args.xlsx)
    datos = semilla.construir()
    args.salida.parent.mkdir(parents=True, exist_ok=True)
    args.salida.write_text(json.dumps(datos, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"OK  {args.salida}")
    print(f"    bloques {len(datos['bloques'])} · camas {len(datos['camas'])} · "
          f"variedades {len(datos['variedades'])} · siembras {len(datos['siembras'])}")
    revisar = [a for a in datos["anomalias"] if a["severidad"] == "revisar"]
    info = [a for a in datos["anomalias"] if a["severidad"] == "informativo"]
    print(f"    anomalías a revisar {len(revisar)} · informativas {len(info)}")
    for a in revisar:
        print(f"      [!] {a['detalle']}")
    for a in semilla.avisos:
        print(f"      (i) {a}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
